import logging
from atlassian import Jira
import requests
import configparser
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
#suppress log WARNING only show ERRORS from atlassian package
logging.getLogger("atlassian").setLevel(logging.ERROR)

class TestPlanNotFoundException(Exception):
    pass

class TestNotFoundException(Exception):
    pass

class PostRequestException(Exception):
    pass

class XrayTestBuilder:
    def __init__(self, jira_instance, base_url, token):
        self._jira = jira_instance
        self._test_issue = None
        self._test_steps = []
        self._base_url = base_url
        self._token = token

    def _send_post_request(self, endpoint, data):
        headers = {
            "Authorization": f"Bearer {self._token}",
            "Content-Type": "application/json"
        }
        response = requests.post(endpoint, json=data, headers=headers)
        if response.status_code != requests.codes.ok:
            raise PostRequestException(f"Request failed. Status code: {response.status_code}, Response: {response.text}")

        return response

    def create_test(self, project_key, summary, test_path, description="example of manual test"):
        steps_data = [{
            "index": index,
            "fields": {
                "action": step["action"],
                "data": step["data"],
                "expected result": step["result"]
            }
        } for index, step in enumerate(self._test_steps, 1)]

        data = {
            "project": {"key": project_key},
            "summary": summary,
            #TODO: Make shure that your Test path matches the custom field - in my case its customfield_12320
            "customfield_12320": test_path,
            "description": description,
            "issuetype": {"name": "Test"},
            #TODO: Make shure that your Test Type "Manual" matches the custom field - in my case its customfield_12310
            "customfield_12310": {"value": "Manual"},
            #TODO: Make shure that your Test Steps matches the custom field - in my case its customfield_12314
            "customfield_12314": {"steps": steps_data},
        }
        self._test_issue = self._jira.issue_create(fields=data)
        jira_test_key = self._test_issue['key']
        logger.info(f"Success! Test created.\n{self._test_issue}\nvisit {self._base_url}/browse/{jira_test_key}")
        return self

    def add_test_step(self, action, data, expected_result):
        self._test_steps.append({
            "action": action,
            "data": data,
            "result": expected_result
        })
        return self

    def _add_test_to_jira_part(self, part, key):
        if not self._test_issue:
            raise TestNotFoundException(f"You need to create a test before updating the {part}.")

        endpoint = f"{self._base_url}/rest/raven/1.0/api/{part}/{key}/test"
        data = {"add": [self._test_issue["key"]]}
        self._send_post_request(endpoint, data)
        return self

    def add_test_to_testplan(self, plan_key):
        return self._add_test_to_jira_part("testplan", plan_key)

    def add_test_to_test_execution(self, execution_key):
        return self._add_test_to_jira_part("testexec", execution_key)

class Configuration:
    def __init__(self, config_path="settings.ini"):
        self._config_path = config_path
        self._load_configurations()

    def _load_configurations(self):
        config = configparser.ConfigParser()
        config.read(self._config_path)
        self.url = config['DEFAULT']['url']
        self.project = config['DEFAULT']['project']
        self.excel_filepath = config['DEFAULT']['excel_filepath']
        self.token = config['DEFAULT']['token']

def load_tests_from_excel(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    tests = []
    current_test = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_name, test_path, description, action, data, expected_result, plan_key, execution_key = row

        # If new test found
        if test_name:
            if current_test:  # If there's an existing test, add it to the tests list
                tests.append(current_test)
            current_test = {
                "name": test_name,
                "path": test_path,
                "description": description,
                "steps": [],
                "plan_key": plan_key,
                "execution_key": execution_key
            }
        current_test["steps"].append({
            "action": action,
            "data": data,
            "expected_result": expected_result
        })

    if current_test:  # Add the last test if exists
        tests.append(current_test)
    return tests

if __name__ == "__main__":
    config = Configuration()
    jira = Jira(url=config.url, token=config.token)
    jira_project = config.project
    
    #Load Test from the excel file
    tests_data = load_tests_from_excel(config.excel_filepath)
    
    #iterate through test data excel file
    for test_data in tests_data:
        manual_test_builder = XrayTestBuilder(jira, config.url, config.token)
        #add steps to the manual test case
        for step in test_data["steps"]:
            manual_test_builder.add_test_step(step["action"], step["data"], step["expected_result"])
        #create test with all steps
        manual_test_builder.create_test(jira_project, test_data["name"], test_data["path"], test_data["description"])
        #add test to the testplan if a test plan is in excel 
        if test_data["plan_key"]:
            manual_test_builder.add_test_to_testplan(test_data["plan_key"])
        #add test to test execution if the test execution is in excel
        if test_data["execution_key"]:
            manual_test_builder.add_test_to_test_execution(test_data["execution_key"])
